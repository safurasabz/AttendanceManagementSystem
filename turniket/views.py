from django.shortcuts import render
import openpyxl, re
from turniket.models import Excel, Staff, Log, Permissions, Shortday
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, F, Sum
from django.db.models import Value as V
from django.http import HttpResponse
import datetime
import csv
import time

# Create your views here.

app_name = "turniket"

def addperm(request):
    data = request.POST.copy()
    ddd = str(data["date"])
    perm = Permissions(name=data["staffname"], date=data["date"], hour1=data["hour1"]+":00", hour2=data["hour2"]+":00",
                       reason=data["reason"], note =data["note"])
    perm.save()
    #print("saved")
    return render(request, 'turniket/index.html', {"data": data})


def index(request):
    staff = Staff.objects.all()
    if "GET" == request.method:
        return render(request, 'turniket/index.html', {"names": Staff.objects.all()})
    else:
        excel_file = request.FILES["excel_file"]
        stopw = time.time()
        print("lalal")
        print("EXCEL file accepted")
        print("excel file is trying to be loaded")
        wb = openpyxl.load_workbook(excel_file)
        print("excel file loaded")
        sheets = wb.sheetnames
        noway = "NODATA"
        cnt = 11
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        excel_data = list()
        print("Begin to delete excel old files")
        Excel.objects.all().delete()
        print("Old excel objects deleted")
        now = datetime.datetime.now()
        print("stopwatch started")
        #print(now)
        threemonthsago = str(now)[:5] + str((int(str(now)[5:7])+9)%12) + str(now)[7:]
        print("3monthsagosetted")
        #print(threemonthsago)
        #print(threemonthsago>"2020-01-30")
        print("OLD log objects filtered will be deleted ")
        Log.objects.filter(date__lt=threemonthsago).delete()
        print("Log objects filtered deleted")
        #excelfileden oxuyub liste yazmaq ve listi dba save etmek
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            if row_data[9] == "" or row_data[0] == "" or row_data[9] is None or row_data[0] is None or row_data[9] == "None" or row_data[0]=="None" :
                continue
            else:
                dataRow = Excel(dat=row_data[0], time=row_data[1], action=row_data[3], name=row_data[9])
                print("row data is selected")
                dataRow.save()
                print("row data is saved to excel db")
                excel_data.append(row_data)
        dateslist = []
        #faylda distinct insanlari iterate edir
        for data in Excel.objects.values('name').distinct():
            print("distinct names in excel file is sorted...")
            print(data)
            #print(data)
            #fayldaki adlarin her birini ayriliqda sechib obyektleri dates-e yigir.
            if len(Staff.objects.filter(name = data['name']))==0:
                person = Staff(name = data['name'], excel_name = data['name'], status=1)
                person.save()


            dates = Excel.objects.filter(name=data['name'])

            #obyektleri yigdiqdan sonra iterate edib list yaradir
            for date in dates:
                dateslist.append(date.dat)
            #distinctlesdhrir dateslist date stringini saxlayir
            dateslist = list(dict.fromkeys(dateslist))
            #her bir date-e gore log db uchun record yigilmaya bashlayir
            #her bir insana aid olan date-leri bir br iterate edirik

            for date in dateslist:
                if len(Log.objects.filter(fullname=data['name'], date=date)) > 0:
                    continue
                logdate = date
                hours = []
                actions = []
                logfullname = data['name']

                logfinishhour = "18:00:00"
                loggraph = "09:00 - 18:00"

                if len(Shortday.objects.filter(date=date)) > 0:
                    loggraph = "09:00 - 17:00"
                    logfinishhour = '17:00:00'

                #her bir insanin bir gun ichinde aid olan saatlari ve actionlari list e yigir ardicil
                for i in Excel.objects.filter(name=data['name'], dat=date):
                    hours.append(i.time)
                    actions.append(i.action)

                logbuild_exit = 0
                #gelme tarixi varsa gecikmeni tez gelmeni ve gelme vaxtini tapir
                if len(hours) > 0 and actions[0] == "вход" and hours[0] != "":
                    logcame = hours[0]
                    hours.remove(logcame)
                    #print(actions[0])
                    del actions[0]
                    if logcame[:2] != "" and int(logcame[:2]) < 9 and logcame!=noway:
                        logearly_came = convertTimeToStr(difference('09:00:00', logcame))
                        loglate_came = "0"
                    elif logcame[:2] != "" and int(logcame[:2]) >= 9 and logcame!=noway:
                        logearly_came = "0"
                        loglate_came = convertTimeToStr(difference(logcame, '09:00:00'))
                else:
                    logcame = noway
                    loglate_came = noway
                    logearly_came = noway
                    logoveralldec = 0
                    logoverallgdec = 0
                    logoverall = noway
                    logoverallg = noway


                if len(hours) > 0 and actions[len(hours)-1] == "выход" and hours[-1] != "":
                    loggone = hours[len(hours)-1]
                    hours.remove(loggone)
                    del actions[len(actions)-1]
                    if loggone[:2] != "" and int(loggone[:2]) >= int(logfinishhour[:2]) and loggone!=noway:
                        loglate_gone = convertTimeToStr(difference(loggone, logfinishhour))
                        logearly_gone = "0"
                    elif loggone[:2] != "" and int(loggone[:2]) < int(logfinishhour[:2]) and loggone!=noway:
                        logearly_gone = convertTimeToStr(difference(logfinishhour, loggone))
                        loglate_gone = "0"
                else:
                    loggone = noway
                    logearly_gone = noway
                    loglate_gone = noway
                    logoverall = noway
                    logoveralldec = 0
                    logoverallgdec = 0
                    logoverallg = noway
               # print(logfullname)
               # print(hours)
               # print(actions)
                actbin = []
                for v in actions:
                    if v == "выход":
                        actbin.append(1)
                    else:
                        actbin.append(0)
                logbuild_exit = sum(actbin)
                log_modalnum=""
                log_modalnumdash = ""
                distractiondec = 0
                distractionstr = ""
                log_hours = ""

                for j in range(len(hours) - 1):
                    if actbin[j] == 1 and actbin[j + 1] == 0:
                        if int(hours[j][:2])<13:
                            left = hours[j]
                        elif int(hours[j][:2])>=14:
                            left = hours[j]
                        else:
                            left = '14:00:00'

                        if int(hours[j+1][:2])< 13:
                            entr = hours[j+1]
                        elif int(hours[j+1][:2])>=14:
                            entr = hours[j+1]
                        else:
                            entr = '13:00:00'
                       # print("entrleftbefor", entr, left)
                        if int(entr[:2])>=int(left[:2]):
                           # print("entrleft", entr, left)
                           if int(hours[j+1][:2])>14 and int(hours[j][:2])<13:
                               distractiondec = distractiondec + difference(entr, left) - difference('14:00:00', '13:00:00')
                           else:
                               distractiondec = distractiondec + difference(entr, left)
                        log_hours = log_hours + hours[j] + "-" + hours[j+1] + " ;              "
                        cnt = cnt + 1
                        log_modalnum = "modal" + str(cnt)
                        log_modalnumdash = "#" + log_modalnum

                if loggone != noway and logcame != noway:
                    if int(loggone[:2]) >=18:
                        last =  "18:00:00"
                    else:
                        last = loggone
                    if int(logcame[:2]) < 9:
                        first = "09:00:00"
                    else:
                        first = logcame
                    if int(loggone[:2])<=13:
                        gonebefbr =  0
                    else:
                        gonebefbr = difference('14:00:00', '13:00:00')
                    logoverall = convertTimeToStr(
                        difference(loggone, logcame) - distractiondec - gonebefbr)
                    logoverallg = convertTimeToStr(
                        difference(last, first) - distractiondec - gonebefbr)

                if len(Permissions.objects.filter(name=data['name'], date=logdate))>0:
                    permission = Permissions.objects.filter(name=data['name'], date=logdate)
                    for perm in permission:
                        if int(perm.hour1[:2]) <= 13 <= int(perm.hour2[:2]):
                            nondist = difference(perm.hour2, perm.hour1) - difference('14:00:00', '13:00:00')
                            distractiondec = distractiondec - nondist
                            logbuild_exit = logbuild_exit - 1

                logoveralldec = convertStrToSec(logoverall)
                logoverallgdec = convertStrToSec(logoverallg)
                distractionstr = convertTimeToStr(distractiondec)
                logovertimedec = logoveralldec - logoverallgdec



                if convertStrToSec(loglate_came) >0 or len(loglate_came)>6 or loglate_came!=noway:
                    loglate_camebin = 1
                    loglate_camedec = convertStrToSec(loglate_came)
                else:
                    loglate_camebin=0
                    loglate_camedec = 0


                #whether the employee has permission for current date
                p = Permissions.objects.filter(name=data['name'], date=date)
                if  len(p)>0:
                    for perm in p:
                        logpermission = str(perm.hour1)[:5] + "-" + str(perm.hour2)[:5]
                        logpermissionreason = perm.reason + " not:" + perm.note
                else:
                    logpermission = "-"
                    logpermissionreason = "-"

                if logearly_gone!=noway:
                    logearly_gonedec = convertStrToSec(logearly_gone)
                else:
                    logearly_gonedec=0

                logdata = Log(
                    date=logdate,
                    fullname=logfullname,
                    graph = loggraph,
                    came=logcame,
                    gone=loggone,
                    early_came=logearly_came,
                    late_came=loglate_came,
                    late_camedec=loglate_camedec,
                    late_camebin = loglate_camebin,
                    early_gone=logearly_gone,
                    early_gonedec = logearly_gonedec,
                    late_gone=loglate_gone,
                    build_exit=logbuild_exit,
                    distraction=distractionstr,
                    overall_hourdec = logoveralldec,
                    overall_ghourdec = logoverallgdec,
                    overall_ghour=logoverallg,
                    overall_hour=logoverall,
                    overtimedec = logovertimedec,
                    permission=logpermission,
                    permissionreason=logpermissionreason,
                    hour_str = log_hours,
                    modalnum = log_modalnum,
                    modalnumdash = log_modalnumdash)
                logdata.save()
                stopend = time.time()
                print("Duration:", stopend - stopw)
        return render(request, 'turniket/index.html', {"excel_data": Log.objects.all(), "names": staff})




def filter(request):
    data = request.POST.copy()
    ddd = str(data["date1"])
    keys = data.keys()
    #print('from filter func',data, keys)
    for temp in Log.objects.filter(fullname=data['fullname'], date=str(data["date1"])):
        if temp.overall_ghour!="0" and temp.overall_hour!="0":
            graphichour = convertStrToSec(temp.overall_ghour)
            totalhour = convertStrToSec(temp.overall_hour)
        else:
            graphichour = convertStrToSec("0 s 0 deq 0 san")
            totalhour = convertStrToSec("0 s 0 deq 0 san")
        late = convertStrToSec(temp.late_came)

    belateint = int(data["belate"])*60
    overtimeint = int(data["overtime"])*60

    #print("filterdatas:", data, keys, "\n")
    #print(belateint, overtimeint)

    if data['fullname']!= '':
        if data['date1']!='':
            if data['date2']!='':
                excel_file = Log.objects.filter(fullname=data['fullname'], date__gte=data["date1"], date__lte=data["date2"], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint) )
            else:
                excel_file = Log.objects.filter(fullname=data['fullname'], date=str(data["date1"]), overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint))
        else:
            excel_file = Log.objects.filter(fullname=data['fullname'], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint) )


    else:
        if data['date1'] != '':
            if data['date2']!='':
                excel_file = Log.objects.filter(date__gte=data["date1"], date__lte=data["date2"], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint))
            else:
                excel_file = Log.objects.filter(date=data['date1'], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint))
        else:
            excel_file = Log.objects.filter(late_camedec__gte=Value(belateint), overtimedec__gte=Value(overtimeint))

    return render(request, 'turniket/index.html', {"excel_data": excel_file, "names": Staff.objects.all()})


def addshortday(request):
    if request.method == "GET":
        return render(request, 'turniket/index2.html')
    else:
        data = request.POST.copy()
        ddd = str(data["date"])
        sd = Shortday(date=data['date'])
        sd.save()
        return render(request, 'turniket/index.html')


def difference(hour1, hour2):
    if hour1 != " " and hour2 != " ":
        #print(hour1, hour2)
        h1 = hour1[:2]
        h2 = int(hour2[:2])
        m1 = hour1[3:5]
        m2 = int(hour2[3:5])
        s1 = int(hour1[6:])
        s2 = int(hour2[6:])
        time1 = int(h1)*3600 + int(m1)*60 + s1
        time2 = h2*3600 + m2*60 + s2
        return time1 - time2
    else:
        return 0

def convertTimeToStr(time):
    #print(time)
    if time != 0:
        h = time//3600
        res = str(h) + " s "
        m = (time - (time//3600)*3600)//60
        res = res + str(m) + " dəq "
        s = time - h*3600 - m*60
        res = res + str(s) + " san "
        return res
    else:
        return "0"

def convertStrToSec(timeline):
    # print('from converting ti dec func: input is :', timeline)
    templist = timeline.split()
    if len(templist)>=5:
        # print(templist)
        h = int(templist[0])
        m = int(templist[2])
        s = int(templist[4])
        decimaltime = h*3600 + m*60 + s
        # print('from converting ti dec func: output is :', decimaltime)
        return decimaltime
    else:
        return 0





def exportreport(request):

    data = request.POST.copy()

    keys = data.keys()

    belateint = int(data["belate"]) * 60
    overtimeint = int(data["overtime"]) * 60
    #print(data)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Hesabat.csv'
    response.write(u'\ufeff'.encode('utf8'))

    writer = csv.writer(response)

    writer.writerow(['Əməkdaş', 'Tarix', 'Giriş', 'Çıxış', 'Gecikmə', 'Tez çıxma'])

    tdata = Log.objects.all().values_list('fullname','date', 'came', 'gone', 'late_came', 'early_gone')
    names = []

    for log in Log.objects.filter(date__gte=data["date1"], date__lte=data["date2"], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint)):
        if log.fullname not in names:
            names.append(log.fullname)

    for name in names:
        objs = Log.objects.filter(fullname = name, date__gte=data["date1"], date__lte=data["date2"], overtimedec__gte=V(overtimeint), late_camedec__gte=V(belateint))
        tdata = objs.values_list('fullname', 'date', 'came', 'gone', 'late_came', 'early_gone')
        for idata in tdata:
            writer.writerow(idata)
        exttuple=[name, ' ', ' ', ' ', convertTimeToStr(objs.aggregate(Sum('late_camedec'))['late_camedec__sum']), convertTimeToStr(objs.aggregate(Sum('early_gonedec'))['early_gonedec__sum'])]
        exttuple = tuple(exttuple)

        writer.writerow(exttuple)

    return response